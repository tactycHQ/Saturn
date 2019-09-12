from openpyxl import load_workbook, workbook
import pandas as pd
from pycel.excelcompiler import ExcelCompiler
import networkx as nx

class Saturn:

    def __init__(self, file):
        self.file = file
        self.excel = ExcelCompiler(file)

    def load_excel(self,data_only=False):
        wb = load_workbook(filename=self.file, data_only=data_only)
        return wb

    def getCells(self, wb):
        all_addrs = []

        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows():
                for cell in row:
                    if cell.value:
                        address = "{}!{}".format(sheet,cell.coordinate)
                        all_addrs.append(address)

        self.all_cells = all_addrs
        return self.all_cells

    def initializePycel(self, all_addrs):
        for address in all_addrs:
            self.excel.evaluate(address)
        # print(self.excel.cell_map)
        # print(self.excel._formula_cells_dict)


    def processGraph(self):
        G = self.excel.dep_graph
        graph_cells=[]

        for cell in G:
            graph_cells.append(cell.address)
        self.graph_cells = graph_cells

        return self.graph_cells

    @property
    def getLabels(self):
        labels=[]

        for address in all_cells:
            if address not in str(self.graph_cells):
                labels.append(address)
                print(self.excel.evaluate(address))
        self.labels=labels

        return self.labels

    def validatePycel(self):
        self.excel.validate_calcs()


if __name__ ==  '__main__':

    xlsname = "../TestModel_v1.xlsx"
    xlc = Saturn(xlsname)
    wb = xlc.load_excel(data_only=False)
    all_cells = xlc.getCells(wb)
    xlc.initializePycel(all_cells)
    graph_addrs = xlc.processGraph()
    print(xlc.getLabels)






    # xlc.validatePycel()








    # computed_cells = xlc.calculateCells(all_cells)

