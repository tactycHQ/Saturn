import logging

logger = logging.getLogger(__name__)
from openpyxl import load_workbook, workbook
from rpnnode import RPNNode, OperatorNode, RangeNode, OperandNode, FunctionNode
from cell import Cell
from tqdm import tqdm
from excellib import *

class Loader:
    """
    Class responsible for injecting an xlsx file and converting the file into Cell objects.
    1. Injects an xlsx file
    2. Uses openpyxl to convert extract formulas and value from each cell
    3. Instantiates Cell objects for each cell extracted
    """

    def __init__(self, file):
        '''
        Initializes a loader object and sets the self.file param to injected file
        @param file: xlsx file
        '''
        self.file = file
        logging.info("Loading excel file...")

        # Load file in read-only mode for faster execution. Need to read twice to extract formulas separately (Is there a better way?)
        self.wb_data_only = load_workbook(filename=self.file, data_only=True, read_only=True)
        logging.info("Values Loaded...")

        self.wb_formulas = load_workbook(filename=self.file, data_only=False, read_only=True)
        logging.info("Formulas Loaded...")

        logging.info("Excel file loaded")

        self.cells = {}
        self.precMap = {}
        self.depMap = {}

        #Parse cells for value
        self.parseCells()

        #Make cells with just RPN for now. AST tree has not been compiled yet
        self.makeCells()

    def parseCells(self):
        '''
        Extraction of values and formulas from each source cell
        @return: Generates a val_dict for values extracted and form_dict for formulas extracted
        '''
        val_dict = {}
        form_dict = {}

        logging.info("Extracting values from cells")
        for sheet in self.wb_data_only.sheetnames:
            for row in self.wb_data_only[sheet].iter_rows():
                for cell in row:
                    if cell.value is not None:
                        address = "{}!{}".format(sheet, cell.coordinate)
                        val_dict[address] = cell.value

        logging.info("Extracting formulas from cells")
        for sheet in self.wb_formulas.sheetnames:
            for row in self.wb_formulas[sheet].iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # logging.debug(cell.coordinate)
                        address = "{}!{}".format(sheet, cell.coordinate)
                        form_dict[address] = cell.value

        logging.info("Retrieved list of {} values".format(len(val_dict)))
        logging.info("Retrieved list of {} formulas".format(len(form_dict)))

        # Test that values extracted match formulas extracted
        assert len(val_dict) == len(form_dict), \
            "Extraction Mismatch - Formulas extracted:{}, Valued extracted:{}".format(len(form_dict), len(val_dict))

        self.val_dict = val_dict
        self.form_dict = form_dict

    def makeCells(self):
        '''
        Wrapper function that instantiates Cell objects for each extracted cell
        @return: Returns self.cells, a list of Cell objects, created
        '''
        for (k, v), (k2, f) in zip(self.val_dict.items(), self.form_dict.items()):
            cell = self.makeCell(k)
        logging.info("--------{} total cell objects created------------".format(len(self.cells)))

        # Now that all cells are created, go ahead and create dependency map
        self.createDepMap()

    def makeCell(self, address):
        '''
        Wrapper function that instantiates 1 Cell object for each extracted cell
        @return: Returns the cell object
        '''
        if address not in self.cells:
            cell = Cell(address)
            logging.info("Not in cellmap, so making cell {}".format(address))
            cell.value = self.val_dict.get(address)
            cell.formula = self.form_dict.get(address)
            logging.info("1 cell object created for {} with value {}".format(address, cell.value))

            #Update master cell dictionary
            self.cells[cell.address] = cell

            # Update master precedent dictionary
            self.precMap.update({address: cell.prec})
            return cell

        else:
            logging.info('-----Found {} in cellmap----'.format(address))
            return self.cells.get(address)

    def createDepMap(self):
        '''
        Creates computed list of dependents from the precedent map
        @return: Updates a dictionary of precedents with format {address: [list of dependent addresses]}
        '''
        for address, precedents in self.precMap.items():
            for prec in precedents:
                self.depMap.setdefault(prec, []).append(address)

    def getCell(self, address):
        '''
        Returns a cell object at a specified address
        @param address: address
        '''
        try:
            return self.cells.get(address)
        except Exception as ex:
            logging.info("No cell found")

    def getvalue(self,address):
        '''
        Gets value of cell
        @param address: address of desired cell's value
        @return: cell value
        '''
        try:
            return self.getCell(address).value
        except:
            logging.info("Empty cell found at {}. Setting value to zero".format(address))
            return 0

    def setvalue(self, newvalue, address):
        '''
        Sets value of a cell to a specified new value
        @param newvalue: New value to be set
        @param address: Cell to be set
        '''

        # Set value of cell object to new value
        cell = self.getCell(address)
        cell.value = newvalue

        # Update master cell list with new cell
        self.cells[address] = cell
        logging.info("Value in cell {} set to {}".format(address, newvalue))

        # Update dependent cells values
        self.updateDepCells(address)

    def setformula(self, newform, address):
        '''
        Sets value of a cell to a specified new value
        @param newvalue: New value to be set
        @param address: Cell to be set
        '''

        # Set value of cell object to new value
        cell = self.getCell(address)
        cell.formula = newform

        #Calculate new formula
        self.calculate(cell)

        logging.info("Formula in cell {} set to {}".format(address, newform))

    def updateDepCells(self, address):
        '''
        Update all dependent cells for a given address
        @param address: Address of source cell that has been changed
        @return:
        '''
        dep_addrs = self.depMap.get(address)


        # First check if there any dependents. No dependents if cell is an output
        if dep_addrs:
            logging.info("Updating dependent cells {}".format(dep_addrs))
            for addrs in dep_addrs:
                dep = self.getCell(addrs)
                self.calculate(dep)
        else:
            logging.info("No more dependent cells found.")
            pass

    def calculate(self, cell):
        '''
        Calculates the formula in a cell using post order traversal of RPN
        @param cell: Cell to be calculated
        @return: Value of calculation
        '''

        logging.info("******** Evaluating cell {} with RPN {}".format(cell.address, cell.rpn))

        tree = cell.rpn
        stack = []

        for node in tree:
            logging.info("Processing node: {}".format(node))

            #Operator Node
            if isinstance(node, OperatorNode):
                if node.token.type == node.token.OP_IN:
                    arg2 = stack.pop()
                    arg1 = stack.pop()
                    op = OP_MAP.get(node.token.value)
                    eval_str = '''{}{}{}'''.format(arg1, op, arg2)
                    result = eval(eval_str)
                    stack.append(result)
                    logging.info("Found operator node with value {}".format(node.token.value))
                else:
                    arg1 = stack.pop()
                    op = OP_MAP.get(node.token.value)
                    eval_str = '''{}{}'''.format(op, arg1)
                    result = eval(eval_str)
                    stack.append(result)

            #Function Node
            elif isinstance(node,FunctionNode):
                if node.num_args:
                    args = stack[-node.num_args:]
                    del stack[-node.num_args:]
                    func = FUNC_MAP.get(node.token.value.strip('('))

                    temp_args=[]
                    for i,a in enumerate(args):
                        temp_args.append(a)

                    arg_str = '{}'.format(tuple(temp_args))
                    logging.info("Argument string is: {}".format(arg_str))
                    eval_str = '{}{}'.format(func, arg_str)
                    logging.info("Evaluate code is: {}".format(eval_str))
                    result = eval(eval_str)
                    stack.append(result)

            #Operand Node - Number type
            elif node.token.subtype == 'NUMBER':
                 stack.append(int(node.token.value))
                 logging.info("Found number node with value {}".format(node.token.value))

            #Operand node - Range type
            else:
                range_tuple = None
                if len(node.rangeadds) > 1:
                    range_val = []
                    for adds in node.rangeadds:
                        slice_= []
                        for add in adds:
                            slice_.append(self.getvalue(add))
                        range_val.append(tuple(slice_))
                        range_tuple = tuple(range_val)

                else:
                    range_val = []
                    adds = node.rangeadds[0][0]
                    range_val = self.getvalue(adds)
                    range_tuple = range_val

                stack.append(range_tuple)

                logging.info("Found range node {} with value {}".format(node.token.value,range_tuple))

            logging.info("Stack is: {}".format(stack))

        #Get final result from remaining value in stack
        assert len(stack) == 1, 'More than 1 remaining value in stack. Recheck the stack algorithm'
        result = stack.pop()

        #Set cell value to new calculated value
        self.setvalue(result,cell.address)


